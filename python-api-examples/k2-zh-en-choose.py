#!/usr/bin/env python3

"""
中英双语ASR选择工具。

该脚本读取两个ASR结果文件（中文ASR和英文ASR），
根据文本语言比例和置信度，智能选择最合适的ASR结果。

输入格式支持（与 k2-asr-short.py 输出格式兼容）：
- 带置信度：<segment_id> <confidence> <asr_text>
- 不带置信度：<segment_id> <asr_text>

输出Excel文件包含以下列：
- segment_id: 音频片段ID（与输入文件一致）
- 中文ASR text: 中文ASR结果的文本
- 中文ASR confidence: 中文ASR结果的置信度
- 中文ASR中文比例: 中文ASR文本中的中文字符比例（0.0-1.0）
- 英文ASR text: 英文ASR结果的文本
- 英文ASR confidence: 英文ASR结果的置信度
- 判断语言结果: 最终判断的语言（中文/英文/中英/无法判断）
- 选择哪路: 最终选择的ASR路径（中文ASR/英文ASR/两路都不选）

判断逻辑（按规则顺序判断）
1. 如果中文ASR的中文比例=0%，且英文ASR置信度大于高阈值，则语言是英文，选英文ASR
2. 如果中文ASR的中文比例=100%，且中文ASR置信度大于高阈值，则语言是中文，选中文ASR
3. 如果中文ASR的中文比例>50%，且中文ASR置信度大于高阈值，则语言是中英，选中文ASR
4. 如果两路置信度都低于低置信度阈值（默认0.5），则语言是无法判断，选择"两路都不选"。 注意，文本为空，不用计算中文比例，并且认为置信度为0。
5. 如果其中有一路大于0.85，则选择该路，语言为该路对应的语言。
6. 其它情况选择更高置信度的一路，语言为该路对应的语言。
使用示例：

python3 ./python-api-examples/k2-zh-en-choose.py \
  --zh-asr /path/to/zh_asr_result.txt \
  --en-asr /path/to/en_asr_result.txt \
  --output /path/to/result.xlsx \
  --confidence-threshold 0.7 \
  --low-confidence-threshold 0.5
"""

import argparse
import re
import sys
from pathlib import Path
from typing import Dict, Tuple, Optional

try:
    import pandas as pd
except ImportError:
    print("错误: 需要安装 pandas 和 openpyxl")
    print("请运行: pip install pandas openpyxl")
    sys.exit(1)


def parse_asr_line(line: str) -> Tuple[str, Optional[float], str]:
    """
    解析ASR结果文件的一行。
    
    支持三种格式：
    1. segment_id confidence text
    2. segment_id confidence (只有置信度，没有文本)
    3. segment_id text (无置信度)
    
    返回:
        (segment_id, confidence, text)
        confidence 为 None 如果不存在
        text 为空字符串如果不存在
    """
    line = line.strip()
    if not line:
        return None, None, None
    
    parts = line.split(None, 2)  # 最多分割成3部分
    
    if len(parts) == 2:
        # 可能是: segment_id confidence 或 segment_id text
        segment_id, part2 = parts
        # 尝试将第二部分解析为浮点数
        try:
            confidence = float(part2)
            # 如果成功解析为浮点数，且值在合理范围内，认为是置信度（没有文本）
            if -1.0 <= confidence <= 1.0:
                return segment_id, confidence, ""
            else:
                # 超出范围，认为是文本（无置信度）
                return segment_id, None, part2
        except ValueError:
            # 无法解析为浮点数，认为是文本（无置信度）
            return segment_id, None, part2
    elif len(parts) == 3:
        # 格式: segment_id confidence text 或 segment_id text_with_spaces
        segment_id, part2, text = parts
        try:
            confidence = float(part2)
            # 如果成功解析为浮点数，且值在合理范围内，认为是置信度
            if -1.0 <= confidence <= 1.0:
                return segment_id, confidence, text
            else:
                # 超出范围，可能是文本的一部分
                return segment_id, None, f"{part2} {text}"
        except ValueError:
            # 无法解析为浮点数，认为是文本的一部分
            return segment_id, None, f"{part2} {text}"
    else:
        # 格式错误
        return None, None, None


def load_asr_results(asr_path: Path) -> Tuple[Dict[str, Tuple[Optional[float], str]], list, int]:
    """
    加载ASR结果文件。
    
    返回:
        (results_dict, segment_id_order, valid_line_count)
        results_dict: 字典，映射 segment_id 到 (confidence, text)
        segment_id_order: segment_id 在文件中出现的顺序列表
        valid_line_count: 有效行数（不包括空行和格式错误的行）
        confidence 为 None 如果不存在
    """
    if not asr_path.is_file():
        raise FileNotFoundError(f"ASR结果文件未找到: {asr_path}")
    
    print(f"正在加载ASR结果文件: {asr_path}")
    results = {}
    segment_id_order = []  # 保持segment_id出现的顺序
    has_confidence = False
    valid_line_count = 0
    
    with open(asr_path, 'r', encoding='utf-8') as f:
        for line_num, line in enumerate(f, 1):
            segment_id, confidence, text = parse_asr_line(line)
            
            if segment_id is None:
                if line.strip():  # 非空行但格式错误
                    print(f"  警告: 第 {line_num} 行格式无效，跳过: {line[:50]}...")
                continue
            
            valid_line_count += 1
            
            if confidence is not None:
                has_confidence = True
            
            # 添加到顺序列表（按出现顺序）
            segment_id_order.append(segment_id)
            results[segment_id] = (confidence, text)
    
    print(f"  已加载 {len(results)} 条ASR结果（有效行数: {valid_line_count}）")
    if has_confidence:
        print(f"  包含置信度信息")
    else:
        print(f"  不包含置信度信息")
    
    return results, segment_id_order, valid_line_count


def calculate_language_ratio(text: str) -> Tuple[float, float]:
    """
    计算文本中的中英文字符比例。
    
    返回:
        (chinese_ratio, english_ratio)
        chinese_ratio: 中文字符比例 (0.0-1.0)
        english_ratio: 英文字符比例 (0.0-1.0)
    """
    if not text or not text.strip():
        return 0.0, 0.0
    
    # 中文字符范围：\u4e00-\u9fff (CJK统一汉字)
    chinese_pattern = re.compile(r'[\u4e00-\u9fff]')
    # 英文字符：a-z, A-Z
    english_pattern = re.compile(r'[a-zA-Z]')
    
    total_chars = 0
    chinese_chars = 0
    english_chars = 0
    
    for char in text:
        if char.strip():  # 忽略空格
            total_chars += 1
            if chinese_pattern.match(char):
                chinese_chars += 1
            elif english_pattern.match(char):
                english_chars += 1
    
    if total_chars == 0:
        return 0.0, 0.0
    
    chinese_ratio = chinese_chars / total_chars
    english_ratio = english_chars / total_chars
    
    return chinese_ratio, english_ratio


def choose_asr_result(
    zh_text: str,
    zh_confidence: Optional[float],
    en_text: str,
    en_confidence: Optional[float],
    confidence_threshold: float = 0.7,
    low_confidence_threshold: float = 0.5
) -> Tuple[str, str]:
    """
    根据文本语言比例和置信度选择最合适的ASR结果。
    
    返回:
        (selected_language, selected_path)
        selected_language: "中文"、"英文"、"中英" 或 "无法判断"
        selected_path: "中文ASR"、"英文ASR" 或 "两路都不选"
    """
    # 处理空文本情况：文本为空时，不计算中文比例，置信度视为0
    zh_text_empty = not zh_text.strip()
    en_text_empty = not en_text.strip()
    
    # 如果文本为空，置信度视为0
    zh_conf_value = 0.0 if zh_text_empty else (zh_confidence if zh_confidence is not None else 0.0)
    en_conf_value = 0.0 if en_text_empty else (en_confidence if en_confidence is not None else 0.0)
    
    # 只计算中文ASR的中文比例（如果文本不为空）
    if not zh_text_empty:
        zh_zh_ratio, zh_en_ratio = calculate_language_ratio(zh_text)
    else:
        zh_zh_ratio = 0.0  # 空文本时比例设为0
    
    # 规则4: 如果两路置信度都低于低置信度阈值，则语言是无法判断，选择"两路都不选"
    # 注意：文本为空时置信度视为0，所以也会被这里捕获
    if zh_conf_value < low_confidence_threshold and en_conf_value < low_confidence_threshold:
        return "无法判断", "两路都不选"
    
    # 规则1: 如果中文ASR的中文比例=0%，且英文ASR置信度大于高阈值，则语言是英文，选英文ASR
    if not zh_text_empty and abs(zh_zh_ratio - 0.0) < 1e-6:  # 中文比例=0%
        if en_conf_value > confidence_threshold:
            return "英文", "英文ASR"
    
    # 规则2: 如果中文ASR的中文比例=100%，且中文ASR置信度大于高阈值，则语言是中文，选中文ASR
    if not zh_text_empty and abs(zh_zh_ratio - 1.0) < 1e-6:  # 中文比例=100%
        if zh_conf_value > confidence_threshold:
            return "中文", "中文ASR"
    
    # 规则3: 如果中文ASR的中文比例>50%，且中文ASR置信度大于高阈值，则语言是中英，选中文ASR
    if not zh_text_empty and zh_zh_ratio > 0.5:
        if zh_conf_value > confidence_threshold:
            return "中英", "中文ASR"
    
    # 规则5: 如果其中有一路大于0.85，则选择该路，语言为该路对应的语言
    if zh_conf_value > 0.85:
        return "中文", "中文ASR"
    if en_conf_value > 0.85:
        return "英文", "英文ASR"
    
    # 规则6: 其它情况选择更高置信度的一路，语言为该路对应的语言
    if zh_conf_value >= en_conf_value:
        return "中文", "中文ASR"
    else:
        return "英文", "英文ASR"


def create_selection_dataframe(
    zh_results: Dict[str, Tuple[Optional[float], str]],
    en_results: Dict[str, Tuple[Optional[float], str]],
    segment_id_order: list,
    confidence_threshold: float = 0.7,
    low_confidence_threshold: float = 0.5
) -> pd.DataFrame:
    """
    创建选择结果数据框。
    
    参数:
        zh_results: 中文ASR结果字典
        en_results: 英文ASR结果字典
        segment_id_order: segment_id的顺序列表
        confidence_threshold: 高置信度阈值
        low_confidence_threshold: 低置信度阈值（低于此值且两路都低时，两路都不选）
    
    返回:
        pandas DataFrame
    """
    data = []
    
    for segment_id in segment_id_order:
        zh_conf, zh_text = zh_results.get(segment_id, (None, ""))
        en_conf, en_text = en_results.get(segment_id, (None, ""))
        
        # 计算中文ASR的中文比例
        zh_zh_ratio, _ = calculate_language_ratio(zh_text)
        
        # 选择最合适的ASR结果
        selected_lang, selected_path = choose_asr_result(
            zh_text, zh_conf, en_text, en_conf, confidence_threshold, low_confidence_threshold
        )
        
        data.append({
            "segment_id": segment_id,
            "中文ASR text": zh_text if zh_text else "",
            "中文ASR confidence": zh_conf if zh_conf is not None else "",
            "中文ASR中文比例": zh_zh_ratio if zh_text.strip() else "",
            "英文ASR text": en_text if en_text else "",
            "英文ASR confidence": en_conf if en_conf is not None else "",
            "判断语言结果": selected_lang,
            "选择哪路": selected_path,
        })
    
    df = pd.DataFrame(data)
    return df


def save_to_excel(df: pd.DataFrame, stats_df: pd.DataFrame, output_path: Path):
    """
    将数据框保存到Excel文件。
    
    参数:
        df: 主数据框
        stats_df: 统计信息数据框
        output_path: 输出文件路径
    """
    print(f"正在保存到Excel文件: {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    from openpyxl.styles import Font, PatternFill, Alignment
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 保存主数据
        df.to_excel(writer, sheet_name='ASR选择结果', index=False)
        
        # 获取工作表以设置列宽
        worksheet = writer.sheets['ASR选择结果']
        
        # 设置列宽
        worksheet.column_dimensions['A'].width = 20  # segment_id
        worksheet.column_dimensions['B'].width = 50  # 中文ASR text
        worksheet.column_dimensions['C'].width = 15  # 中文ASR confidence
        worksheet.column_dimensions['D'].width = 15  # 中文ASR中文比例
        worksheet.column_dimensions['E'].width = 50  # 英文ASR text
        worksheet.column_dimensions['F'].width = 15  # 英文ASR confidence
        worksheet.column_dimensions['G'].width = 15  # 判断语言结果
        worksheet.column_dimensions['H'].width = 15  # 选择哪路
        
        # 设置标题行样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置文本列的对齐方式
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            # segment_id 列
            row[0].alignment = Alignment(horizontal="left", vertical="top")
            # text 列
            row[1].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            row[4].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            # confidence 列
            if row[2].value:
                row[2].alignment = Alignment(horizontal="center", vertical="top")
            if row[5].value:
                row[5].alignment = Alignment(horizontal="center", vertical="top")
            # 中文比例列
            if row[3].value:
                row[3].alignment = Alignment(horizontal="center", vertical="top")
            # 判断语言结果列
            row[6].alignment = Alignment(horizontal="center", vertical="top")
            # 选择哪路列
            row[7].alignment = Alignment(horizontal="center", vertical="top")
        
        # 保存统计信息
        stats_df.to_excel(writer, sheet_name='统计信息', index=False)
        
        # 设置统计信息工作表样式
        stats_worksheet = writer.sheets['统计信息']
        stats_worksheet.column_dimensions['A'].width = 30  # 统计项
        stats_worksheet.column_dimensions['B'].width = 15  # 数值
        stats_worksheet.column_dimensions['C'].width = 15  # 百分比
        
        # 设置标题行样式
        for cell in stats_worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置数据行样式
        for row in stats_worksheet.iter_rows(min_row=2, max_row=stats_worksheet.max_row):
            # 统计项列
            row[0].alignment = Alignment(horizontal="left", vertical="center")
            # 数值列
            if row[1].value:
                row[1].alignment = Alignment(horizontal="right", vertical="center")
            # 百分比列
            if row[2].value:
                row[2].alignment = Alignment(horizontal="right", vertical="center")
            
            # 如果是标题行（包含"==="），加粗
            if row[0].value and "===" in str(row[0].value):
                row[0].font = Font(bold=True)
    
    print(f"  成功保存 {len(df)} 条记录到 {output_path}")
    print(f"  统计信息已保存到 '统计信息' sheet")


def create_statistics_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    创建统计信息数据框。
    
    返回:
        pandas DataFrame包含统计信息
    """
    total = len(df)
    stats_data = []
    
    # 总记录数
    stats_data.append({"统计项": "总记录数", "数值": total, "百分比": ""})
    stats_data.append({"统计项": "", "数值": "", "百分比": ""})  # 空行
    
    # 统计语言选择分布
    if '判断语言结果' in df.columns:
        stats_data.append({"统计项": "=== 语言选择分布 ===", "数值": "", "百分比": ""})
        lang_counts = df['判断语言结果'].value_counts()
        for lang, count in lang_counts.items():
            percentage = 100.0 * count / total
            stats_data.append({"统计项": lang, "数值": count, "百分比": f"{percentage:.1f}%"})
        stats_data.append({"统计项": "", "数值": "", "百分比": ""})  # 空行
    
    # 统计路径选择分布
    if '选择哪路' in df.columns:
        stats_data.append({"统计项": "=== 路径选择分布 ===", "数值": "", "百分比": ""})
        path_counts = df['选择哪路'].value_counts()
        for path, count in path_counts.items():
            percentage = 100.0 * count / total
            stats_data.append({"统计项": path, "数值": count, "百分比": f"{percentage:.1f}%"})
        stats_data.append({"统计项": "", "数值": "", "百分比": ""})  # 空行
    
    # 统计平均置信度
    stats_data.append({"统计项": "=== 置信度统计 ===", "数值": "", "百分比": ""})
    zh_conf_values = df['中文ASR confidence'].replace('', None).dropna()
    en_conf_values = df['英文ASR confidence'].replace('', None).dropna()
    
    if len(zh_conf_values) > 0:
        avg_zh_conf = zh_conf_values.mean()
        stats_data.append({"统计项": "中文ASR平均置信度", "数值": f"{avg_zh_conf:.4f}", "百分比": ""})
    
    if len(en_conf_values) > 0:
        avg_en_conf = en_conf_values.mean()
        stats_data.append({"统计项": "英文ASR平均置信度", "数值": f"{avg_en_conf:.4f}", "百分比": ""})
    
    return pd.DataFrame(stats_data)


def print_statistics(df: pd.DataFrame):
    """
    打印统计信息。
    """
    total = len(df)
    
    print(f"\n=== 统计信息 ===")
    print(f"总记录数: {total}")
    
    # 统计语言选择分布
    if '判断语言结果' in df.columns:
        lang_counts = df['判断语言结果'].value_counts()
        print(f"\n=== 语言选择分布 ===")
        for lang, count in lang_counts.items():
            percentage = 100.0 * count / total
            print(f"  {lang}: {count} ({percentage:.1f}%)")
    
    # 统计路径选择分布
    if '选择哪路' in df.columns:
        path_counts = df['选择哪路'].value_counts()
        print(f"\n=== 路径选择分布 ===")
        for path, count in path_counts.items():
            percentage = 100.0 * count / total
            print(f"  {path}: {count} ({percentage:.1f}%)")
    
    # 统计平均置信度
    zh_conf_values = df['中文ASR confidence'].replace('', None).dropna()
    en_conf_values = df['英文ASR confidence'].replace('', None).dropna()
    
    if len(zh_conf_values) > 0:
        avg_zh_conf = zh_conf_values.mean()
        print(f"\n中文ASR平均置信度: {avg_zh_conf:.4f}")
    
    if len(en_conf_values) > 0:
        avg_en_conf = en_conf_values.mean()
        print(f"英文ASR平均置信度: {avg_en_conf:.4f}")


def main():
    parser = argparse.ArgumentParser(
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        description="中英双语ASR选择工具"
    )
    
    parser.add_argument(
        "--zh-asr",
        type=str,
        required=True,
        help="中文ASR结果文件路径"
    )
    
    parser.add_argument(
        "--en-asr",
        type=str,
        required=True,
        help="英文ASR结果文件路径"
    )
    
    parser.add_argument(
        "--output",
        type=str,
        required=True,
        help="输出Excel文件路径"
    )
    
    parser.add_argument(
        "--confidence-threshold",
        type=float,
        default=0.7,
        help="高置信度阈值（默认: 0.7）"
    )
    
    parser.add_argument(
        "--low-confidence-threshold",
        type=float,
        default=0.5,
        help="低置信度阈值，当两路置信度都低于此值时，两路都不选（默认: 0.5）"
    )
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("中英双语ASR选择工具")
    print("=" * 60)
    
    zh_asr_path = Path(args.zh_asr)
    en_asr_path = Path(args.en_asr)
    output_path = Path(args.output)
    
    print(f"\n[1/3] 加载ASR结果文件...")
    zh_results, zh_order, zh_count = load_asr_results(zh_asr_path)
    en_results, en_order, en_count = load_asr_results(en_asr_path)
    
    # 检查两个文件的行数是否一致
    if zh_count != en_count:
        print(f"  警告: 两个输入文件的行数不一致")
        print(f"    中文ASR文件: {zh_count} 行")
        print(f"    英文ASR文件: {en_count} 行")
        print(f"  将只处理共同的segment_id")
    
    # 检查segment_id是否一致
    zh_ids = set(zh_results.keys())
    en_ids = set(en_results.keys())
    common_ids = sorted(zh_ids & en_ids)
    
    if not common_ids:
        raise ValueError("中文ASR和英文ASR结果中没有共同的segment_id！")
    
    # 使用第一个文件的顺序（或共同ID的顺序）
    if zh_order == en_order:
        segment_id_order = zh_order
        print(f"  两个文件segment_id顺序一致: 是")
    else:
        # 使用第一个文件的顺序，但只保留共同的ID
        segment_id_order = [sid for sid in zh_order if sid in common_ids]
        print(f"  两个文件segment_id顺序不一致，使用中文ASR文件的顺序")
    
    print(f"  共同segment_id数量: {len(common_ids)}")
    print(f"  高置信度阈值: {args.confidence_threshold}")
    print(f"  低置信度阈值: {args.low_confidence_threshold}")
    
    print(f"\n[2/3] 创建选择结果数据框...")
    df = create_selection_dataframe(
        zh_results, en_results, segment_id_order, 
        args.confidence_threshold, args.low_confidence_threshold
    )
    
    print(f"\n[3/4] 生成统计信息...")
    stats_df = create_statistics_dataframe(df)
    
    print(f"\n[4/4] 保存到Excel文件...")
    save_to_excel(df, stats_df, output_path)
    
    # 打印统计信息
    print_statistics(df)
    
    print(f"\n完成！结果已保存到: {output_path}")


if __name__ == "__main__":
    main()

