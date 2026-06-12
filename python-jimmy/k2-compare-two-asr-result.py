#!/usr/bin/env python3

"""
Compare two ASR result files and write an Excel report.

This script reads two ASR result files, checks that segment IDs are aligned,
and exports a side-by-side comparison sheet.

Examples:
python3 ./python-jimmy/k2-compare-two-asr-result.py \
  --asr1 /path/to/asr_result1.txt \
  --asr2 /path/to/asr_result2.txt \
  --output /path/to/comparison.xlsx

python3 ./python-jimmy/k2-compare-two-asr-result.py \
  --asr1 /path/to/asr_result1.txt \
  --asr2 /path/to/asr_result2.txt \
  --output /path/to/comparison.xlsx \
  --name1 "Model A" \
  --name2 "Model B"

Notes:
- Input files must have the same number of lines
- Segment IDs must match in the same order
- Supported input formats:
  <segment_id> <text>
  <segment_id> <confidence> <text>
"""

import argparse
import sys
from pathlib import Path
from typing import Dict, Tuple, Optional

try:
    import pandas as pd
except ImportError:
    print("错误: 需要安装 pandas 和 openpyxl")
    print("请运行: pip install pandas openpyxl")
    sys.exit(1)


def parse_asr_line(line: str) -> Tuple[str, Optional[float], str]:
    """
    解析ASR结果文件的一行。
    
    支持三种格式：
    1. segment_id confidence text
    2. segment_id confidence (只有置信度，没有文本)
    3. segment_id text (无置信度)
    
    返回:
        (segment_id, confidence, text)
        confidence 为 None 如果不存在
        text 为空字符串如果不存在
    """
    line = line.strip()
    if not line:
        return None, None, None
    
    parts = line.split(None, 2)  # 最多分割成3部分
    
    if len(parts) == 2:
        # 可能是: segment_id confidence 或 segment_id text
        segment_id, part2 = parts
        # 尝试将第二部分解析为浮点数
        try:
            confidence = float(part2)
            # 如果成功解析为浮点数，且值在合理范围内，认为是置信度（没有文本）
            if -1.0 <= confidence <= 1.0:
                return segment_id, confidence, ""
            else:
                # 超出范围，认为是文本（无置信度）
                return segment_id, None, part2
        except ValueError:
            # 无法解析为浮点数，认为是文本（无置信度）
            return segment_id, None, part2
    elif len(parts) == 3:
        # 格式: segment_id confidence text 或 segment_id text_with_spaces
        segment_id, part2, text = parts
        try:
            confidence = float(part2)
            # 如果成功解析为浮点数，且值在合理范围内，认为是置信度
            if -1.0 <= confidence <= 1.0:
                return segment_id, confidence, text
            else:
                # 超出范围，可能是文本的一部分
                return segment_id, None, f"{part2} {text}"
        except ValueError:
            # 无法解析为浮点数，认为是文本的一部分
            return segment_id, None, f"{part2} {text}"
    else:
        # 格式错误
        return None, None, None


def load_asr_results(asr_path: Path) -> Tuple[Dict[str, Tuple[Optional[float], str]], list, int]:
    """
    加载ASR结果文件。
    
    返回:
        (results_dict, segment_id_order, valid_line_count)
        results_dict: 字典，映射 segment_id 到 (confidence, text)
        segment_id_order: segment_id 在文件中出现的顺序列表
        valid_line_count: 有效行数（不包括空行和格式错误的行）
        confidence 为 None 如果不存在
    """
    if not asr_path.is_file():
        raise FileNotFoundError(f"ASR结果文件未找到: {asr_path}")
    
    print(f"正在加载ASR结果文件: {asr_path}")
    results = {}
    segment_id_order = []  # 保持segment_id出现的顺序
    has_confidence = False
    valid_line_count = 0
    
    with open(asr_path, 'r', encoding='utf-8') as f:
        for line_num, line in enumerate(f, 1):
            segment_id, confidence, text = parse_asr_line(line)
            
            if segment_id is None:
                if line.strip():  # 非空行但格式错误
                    print(f"  警告: 第 {line_num} 行格式无效，跳过: {line[:50]}...")
                continue
            
            valid_line_count += 1
            
            if confidence is not None:
                has_confidence = True
            
            # 添加到顺序列表（按出现顺序）
            segment_id_order.append(segment_id)
            results[segment_id] = (confidence, text)
    
    print(f"  已加载 {len(results)} 条ASR结果（有效行数: {valid_line_count}）")
    if has_confidence:
        print(f"  包含置信度信息")
    else:
        print(f"  不包含置信度信息")
    
    return results, segment_id_order, valid_line_count


def create_comparison_dataframe(
    results1: Dict[str, Tuple[Optional[float], str]],
    results2: Dict[str, Tuple[Optional[float], str]],
    segment_id_order: list,
    name1: str = "ASR1",
    name2: str = "ASR2"
) -> pd.DataFrame:
    """
    创建比较数据框。
    
    参数:
        results1: 第一个ASR结果字典
        results2: 第二个ASR结果字典
        segment_id_order: segment_id的顺序列表（两个文件顺序一致）
        name1: 第一个ASR结果的名称
        name2: 第二个ASR结果的名称
    
    返回:
        pandas DataFrame，按照输入文件的顺序排列
    """
    # 由于两个文件的segment_id顺序完全一致，直接使用顺序列表
    all_segment_ids = segment_id_order
    
    data = []
    for segment_id in all_segment_ids:
        conf1, text1 = results1.get(segment_id, (None, ""))
        conf2, text2 = results2.get(segment_id, (None, ""))
        
        # 计算置信度差异
        conf1_val = conf1 if conf1 is not None else None
        conf2_val = conf2 if conf2 is not None else None
        
        confidence_diff = ""
        diff_level = ""
        
        if conf1_val is not None and conf2_val is not None:
            # 两个都有置信度，计算差异
            confidence_diff = conf1_val - conf2_val
            # 分类差异度
            if confidence_diff < 0:
                diff_level = "错"
            elif confidence_diff < 0.1:
                diff_level = "低"
            elif confidence_diff < 0.2:
                diff_level = "中"
            else:
                diff_level = "高"
        elif conf1_val is not None or conf2_val is not None:
            # 只有一个有置信度，无法计算差异
            diff_level = "无法比较"
        
        data.append({
            "segment_id": segment_id,
            f"text_{name1}": text1 if text1 else "",
            f"confidence_{name1}": conf1_val if conf1_val is not None else "",
            f"text_{name2}": text2 if text2 else "",
            f"confidence_{name2}": conf2_val if conf2_val is not None else "",
            "confidence_diff": confidence_diff if confidence_diff != "" else "",
            "diff_level": diff_level if diff_level else "",
        })
    
    df = pd.DataFrame(data)
    return df


def save_to_excel(df: pd.DataFrame, output_path: Path):
    """
    将数据框保存到Excel文件。
    """
    print(f"正在保存到Excel文件: {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='ASR比较', index=False)
        
        # 获取工作表以设置列宽
        worksheet = writer.sheets['ASR比较']
        
        # 设置列宽
        worksheet.column_dimensions['A'].width = 20  # segment_id
        worksheet.column_dimensions['B'].width = 50  # text1
        worksheet.column_dimensions['C'].width = 12  # confidence1
        worksheet.column_dimensions['D'].width = 50  # text2
        worksheet.column_dimensions['E'].width = 12  # confidence2
        worksheet.column_dimensions['F'].width = 15  # confidence_diff
        worksheet.column_dimensions['G'].width = 12  # diff_level
        
        # 设置标题行样式（如果需要）
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置文本列的对齐方式
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            # segment_id 列
            row[0].alignment = Alignment(horizontal="left", vertical="top")
            # text 列
            row[1].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            row[3].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            # confidence 列
            if row[2].value:
                row[2].alignment = Alignment(horizontal="center", vertical="top")
            if row[4].value:
                row[4].alignment = Alignment(horizontal="center", vertical="top")
            # confidence_diff 列
            if row[5].value:
                row[5].alignment = Alignment(horizontal="center", vertical="top")
            # diff_level 列
            if row[6].value:
                row[6].alignment = Alignment(horizontal="center", vertical="top")
    
    print(f"  成功保存 {len(df)} 条记录到 {output_path}")


def print_statistics(df: pd.DataFrame):
    """
    打印统计信息。
    """
    total = len(df)
    
    print(f"\n=== 统计信息 ===")
    print(f"总记录数: {total}")
    
    # 如果包含置信度信息，计算平均置信度
    conf_cols = [col for col in df.columns if col.startswith('confidence_')]
    
    if len(conf_cols) >= 1:
        conf1_values = df[conf_cols[0]].replace('', None).dropna()
        if len(conf1_values) > 0:
            avg_conf1 = conf1_values.mean()
            print(f"\n第一个ASR结果平均置信度: {avg_conf1:.4f}")
    
    if len(conf_cols) >= 2:
        conf2_values = df[conf_cols[1]].replace('', None).dropna()
        if len(conf2_values) > 0:
            avg_conf2 = conf2_values.mean()
            print(f"第二个ASR结果平均置信度: {avg_conf2:.4f}")
    
    # 统计差异度分布
    if 'diff_level' in df.columns:
        print(f"\n=== 差异度分布统计 ===")
        diff_levels = df['diff_level'].value_counts()
        
        total_comparable = len(df[df['diff_level'] != ''])
        if total_comparable > 0:
            print(f"可比较记录数: {total_comparable}")
            
            for level in ['高', '中', '低', '错', '无法比较']:
                if level in diff_levels:
                    count = diff_levels[level]
                    percentage = 100.0 * count / total_comparable
                    print(f"  {level}: {count} ({percentage:.1f}%)")
            
            # 计算平均差异度
            if 'confidence_diff' in df.columns:
                # 将空字符串转换为None，然后过滤掉
                diff_values = df['confidence_diff'].replace('', None).dropna()
                if len(diff_values) > 0:
                    avg_diff = diff_values.mean()
                    print(f"\n平均置信度差异: {avg_diff:.4f}")
                    print(f"最大差异: {diff_values.max():.4f}")
                    print(f"最小差异: {diff_values.min():.4f}")


def main():
    parser = argparse.ArgumentParser(
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        description="比较两个ASR结果文件并合并到Excel"
    )
    
    parser.add_argument(
        "--asr1",
        type=str,
        required=True,
        help="第一个ASR结果文件路径"
    )
    
    parser.add_argument(
        "--asr2",
        type=str,
        required=True,
        help="第二个ASR结果文件路径"
    )
    
    parser.add_argument(
        "--output",
        type=str,
        required=True,
        help="输出Excel文件路径"
    )
    
    parser.add_argument(
        "--name1",
        type=str,
        default="ASR1",
        help="第一个ASR结果的名称（用于Excel列名）"
    )
    
    parser.add_argument(
        "--name2",
        type=str,
        default="ASR2",
        help="第二个ASR结果的名称（用于Excel列名）"
    )
    
    args = parser.parse_args()
    
    # 加载两个ASR结果文件
    print("=" * 60)
    print("ASR结果比较工具")
    print("=" * 60)
    
    asr1_path = Path(args.asr1)
    asr2_path = Path(args.asr2)
    output_path = Path(args.output)
    
    print(f"\n[1/3] 加载ASR结果文件...")
    results1, order1, count1 = load_asr_results(asr1_path)
    results2, order2, count2 = load_asr_results(asr2_path)
    
    # 检查两个文件的行数是否一致
    if count1 != count2:
        raise ValueError(
            f"两个输入文件的行数不一致！\n"
            f"  第一个文件 ({asr1_path}): {count1} 行\n"
            f"  第二个文件 ({asr2_path}): {count2} 行\n"
            f"请确保两个文件的行数完全一致。"
        )
    
    # 检查segment_id顺序是否一致（可选，但有助于发现问题）
    if order1 != order2:
        print(f"  警告: 两个文件的segment_id顺序不一致，将使用第一个文件的顺序")
        # 检查segment_id集合是否一致
        if set(order1) != set(order2):
            raise ValueError(
                f"两个输入文件的segment_id集合不一致！\n"
                f"  第一个文件独有的segment_id: {set(order1) - set(order2)}\n"
                f"  第二个文件独有的segment_id: {set(order2) - set(order1)}\n"
                f"请确保两个文件的segment_id值和顺序完全一致。"
            )
    else:
        print(f"  两个文件行数一致: {count1} 行")
        print(f"  两个文件segment_id顺序一致: 是")
    
    print(f"\n[2/3] 创建比较数据框...")
    # 使用第一个文件的顺序（两个文件顺序一致）
    df = create_comparison_dataframe(results1, results2, order1, args.name1, args.name2)
    
    print(f"\n[3/3] 保存到Excel文件...")
    save_to_excel(df, output_path)
    
    # 打印统计信息
    print_statistics(df)
    
    print(f"\n完成！结果已保存到: {output_path}")


if __name__ == "__main__":
    main()
